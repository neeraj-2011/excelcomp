import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image  # For image chart if needed
import matplotlib.pyplot as plt
import numpy as np
import io
from openpyxl.chart import LineChart, Reference  # For creating charts directly in Excel

# Set the folder containing the reports and output file name
input_folder = "./input_reports/"  # Replace with your folder path
output_file = 'consolidated_report.xlsx'

# Get the list of Excel files in the folder
excel_files = sorted(glob.glob(input_folder + "*.xlsx"))  # Ensure files are sorted by name

# Initialize a DataFrame for Report 1 to use as the base order
base_df = pd.DataFrame()

# Initialize an empty DataFrame to store the merged data
merged_df = pd.DataFrame()

# Loop through each file and merge data
for i, file in enumerate(excel_files, start=1):
    # Read the file into a DataFrame and rename columns explicitly
    df = pd.read_excel(file, header=0)  # Assuming the first row has headers
    df.columns = ['Transactions', f'Report {i} time(90%)']  # Force consistent column names
    
    if i == 1:
        # Use the first report as the base for order
        base_df = df
    
    # Merge with the main DataFrame
    if merged_df.empty:
        merged_df = df
    else:
        merged_df = pd.merge(merged_df, df, on='Transactions', how='outer')

# Sort by the base DataFrame's order
merged_df = base_df[['Transactions']].merge(merged_df, on='Transactions', how='left')

# Replace NaN with empty strings if needed
merged_df.fillna("", inplace=True)

# Change column names for reports to R1, R2, ..., R5
report_columns = [f'Report {i} time(90%)' for i in range(1, len(excel_files) + 1)]
renamed_columns = {f'Report {i} time(90%)': f'R{i}' for i in range(1, len(excel_files) + 1)}
merged_df.rename(columns=renamed_columns, inplace=True)

# Calculate the variance between consecutive reports
variance_columns = []

# Calculate R5 Vs R1, R5 Vs R2, ..., R5 Vs R4
if 'R5' in merged_df.columns:
    for i in range(1, 5):
        report_column = f'R{i}'
        variance_column = f'R5 Vs R{i}'
        merged_df[variance_column] = ((merged_df['R5'] - merged_df[report_column]) / merged_df['R5']) * 100
        variance_columns.append(variance_column)

# Calculate R4 Vs R3, R3 Vs R2, R2 Vs R1
consecutive_variance = [('R4', 'R3'), ('R3', 'R2'), ('R2', 'R1')]
for r1, r2 in consecutive_variance:
    variance_column = f'{r1} Vs {r2}'
    merged_df[variance_column] = ((merged_df[r1] - merged_df[r2]) / merged_df[r1]) * 100
    variance_columns.append(variance_column)

# Reorder the columns to include the new variance columns
columns_order = ['Transactions'] + [f'R{i}' for i in range(1, len(excel_files) + 1)] + variance_columns
merged_df = merged_df[columns_order]

# Write the renamed columns to the output file
wb = load_workbook(output_file)
ws1 = wb.active
ws1.title = "Data"  # Name first sheet as "Data"
merged_df.to_excel(output_file, index=False, sheet_name="Data")

# Apply Font Color Formatting
ws1 = wb['Data']  # Reload the first sheet
# Define font styles and arrow marks
green_font = Font(color="00b300", bold=True)   # Green
orange_font = Font(color="FFA500", bold=True) # Orange
red_font = Font(color="FF0000", bold=True)    # Red
black_font = Font(color="000000", bold=True)  # Black (for variance columns)

up_arrow = "% ↑"
down_arrow = "% ↓"
no_change = "% →"

# Define the light blue fill for the header row
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Define the border style
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Apply the light blue fill and borders to the header row
for cell in ws1[1]:  # First row (header)
    cell.fill = light_blue_fill
    cell.border = thin_border

# Apply formatting logic to R1, R2, R3, R4, R5
for col in ['R1', 'R2', 'R3', 'R4', 'R5']:
    if col in merged_df.columns:
        col_index = merged_df.columns.get_loc(col) + 1  # Column index for Excel (1-based)
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=col_index, max_col=col_index):
            for cell in row:
                if isinstance(cell.value, (int, float)):  # Ensure it's a number
                    if cell.value >= 2:
                        cell.font = red_font
                    elif 1.8 <= cell.value < 2:
                        cell.font = orange_font
                    else:
                        cell.font = green_font
                # Apply border to the cell
                cell.border = thin_border

# Loop through the variance columns to apply formatting, arrow marks, and borders
for variance_column in variance_columns:
    col_index = merged_df.columns.get_loc(variance_column) + 1  # Column index for Excel (1-based)
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=col_index, max_col=col_index):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Ensure it's a number
                formatted_value = f"{cell.value:.2f}"  # Limit to 2 decimal places
                if cell.value > 0:
                    cell.value = f"{formatted_value} {up_arrow}"
                    cell.font = red_font
                elif cell.value < 0:
                    cell.value = f"{formatted_value} {down_arrow}"
                    cell.font = green_font
                else:
                    cell.value = f"{formatted_value} {no_change}"
                    cell.font = black_font
            # Apply border to the cell
            cell.border = thin_border

# Apply borders to all cells that contain data
for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    for cell in row:
        if cell.value is not None:  # Check if the cell contains a value
            cell.border = thin_border

# --- Create a second sheet for graphs ---
ws2 = wb.create_sheet(title="Graphs")  # Create a new sheet for graphs

# Create charts for R5 vs other reports
chart_start_row = 2  # Starting row in the chart sheet

for i in range(1, 5):
    variance_column = f'R5 Vs R{i}'
    transactions = merged_df['Transactions']
    variances = merged_df[variance_column]

    # Create a line chart
    chart = LineChart()
    chart.title = f"Graph of {variance_column}"
    chart.style = 13  # A predefined style for the chart
    chart.x_axis.title = 'Transactions'
    chart.y_axis.title = 'Variance (%)'

    # Add data to chart
    data = Reference(ws1, min_col=merged_df.columns.get_loc(variance_column) + 2, min_row=2, max_row=len(merged_df) + 1)
    categories = Reference(ws1, min_col=1, min_row=2, max_row=len(merged_df) + 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    # Place the chart in the second sheet
    ws2.add_chart(chart, f'A{chart_start_row}')
    
    # Update row for next chart
    chart_start_row += 20  # Adjust the spacing for the next chart

# Save the workbook with both the data and charts
wb.save(output_file)

print(f"Consolidated report with data and graphs saved to {output_file}")
