import pandas as pd
import glob
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image  # <-- Add this import
import matplotlib.pyplot as plt
import numpy as np
import io

# Set the folder containing the reports and output file name
input_folder = "./input_reports/"  # Replace with your folder path
output_file = 'consolidated_report.xlsx'

# Get the list of Excel files in the folder
excel_files = sorted(glob.glob(input_folder + "*.xlsx"))  # Ensure files are sorted by name

# Initialize a DataFrame for Report 1 to use as the base order
base_df = pd.DataFrame()

# Initialize an empty DataFrame to store the merged data
merged_df = pd.DataFrame()

# Loop through each file and merge data
for i, file in enumerate(excel_files, start=1):
    # Read the file into a DataFrame and rename columns explicitly
    df = pd.read_excel(file, header=0)  # Assuming the first row has headers
    df.columns = ['Transactions', f'Report {i} time(90%)']  # Force consistent column names
    
    if i == 1:
        # Use the first report as the base for order
        base_df = df
    
    # Merge with the main DataFrame
    if merged_df.empty:
        merged_df = df
    else:
        merged_df = pd.merge(merged_df, df, on='Transactions', how='outer')

# Sort by the base DataFrame's order
merged_df = base_df[['Transactions']].merge(merged_df, on='Transactions', how='left')

# Replace NaN with empty strings if needed
merged_df.fillna("", inplace=True)

# Change column names for reports to R1, R2, ..., R5
report_columns = [f'Report {i} time(90%)' for i in range(1, len(excel_files) + 1)]
renamed_columns = {f'Report {i} time(90%)': f'R{i}' for i in range(1, len(excel_files) + 1)}
merged_df.rename(columns=renamed_columns, inplace=True)

# Calculate the variance between consecutive reports
variance_columns = []

# Calculate R5 Vs R1, R5 Vs R2, ..., R5 Vs R4
if 'R5' in merged_df.columns:
    for i in range(1, 5):
        report_column = f'R{i}'
        variance_column = f'R5 Vs R{i}'
        merged_df[variance_column] = ((merged_df['R5'] - merged_df[report_column]) / merged_df['R5']) * 100
        variance_columns.append(variance_column)

# Calculate R4 Vs R3, R3 Vs R2, R2 Vs R1
consecutive_variance = [('R4', 'R3'), ('R3', 'R2'), ('R2', 'R1')]
for r1, r2 in consecutive_variance:
    variance_column = f'{r1} Vs {r2}'
    merged_df[variance_column] = ((merged_df[r1] - merged_df[r2]) / merged_df[r1]) * 100
    variance_columns.append(variance_column)

# Reorder the columns to include the new variance columns
columns_order = ['Transactions'] + [f'R{i}' for i in range(1, len(excel_files) + 1)] + variance_columns
merged_df = merged_df[columns_order]

# --- Create the workbook and separate sheets ---
wb = Workbook()
ws_table = wb.active
ws_table.title = "Table"

# Write the merged DataFrame to the first sheet (Table sheet)
for r in dataframe_to_rows(merged_df, index=False, header=True):
    ws_table.append(r)

# Apply Font Color Formatting to Table Sheet
# Apply the light blue fill and borders to the header row
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for cell in ws_table[1]:  # First row (header)
    cell.fill = light_blue_fill
    cell.border = thin_border

# Apply formatting logic to R1, R2, R3, R4, R5 in the Table sheet
for col in ['R1', 'R2', 'R3', 'R4', 'R5']:
    if col in merged_df.columns:
        col_index = merged_df.columns.get_loc(col) + 1  # Column index for Excel (1-based)
        for row in ws_table.iter_rows(min_row=2, max_row=ws_table.max_row, min_col=col_index, max_col=col_index):
            for cell in row:
                if isinstance(cell.value, (int, float)):  # Ensure it's a number
                    if cell.value >= 2:
                        cell.font = Font(color="FF0000", bold=True)  # Red
                    elif 1.8 <= cell.value < 2:
                        cell.font = Font(color="FFA500", bold=True)  # Orange
                    else:
                        cell.font = Font(color="00b300", bold=True)  # Green
                # Apply border to the cell
                cell.border = thin_border

# Loop through the variance columns to apply formatting, arrow marks, and borders in the Table sheet
for variance_column in variance_columns:
    col_index = merged_df.columns.get_loc(variance_column) + 1  # Column index for Excel (1-based)
    for row in ws_table.iter_rows(min_row=2, max_row=ws_table.max_row, min_col=col_index, max_col=col_index):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Ensure it's a number
                formatted_value = f"{cell.value:.2f}"  # Limit to 2 decimal places
                if cell.value > 0:
                    cell.value = f"{formatted_value} % ↑"
                    cell.font = Font(color="FF0000", bold=True)  # Red
                elif cell.value < 0:
                    cell.value = f"{formatted_value} % ↓"
                    cell.font = Font(color="00b300", bold=True)  # Green
                else:
                    cell.value = f"{formatted_value} % →"
                    cell.font = Font(color="000000", bold=True)  # Black
            # Apply border to the cell
            cell.border = thin_border

# --- Create the second sheet for graphs ---
ws_graphs = wb.create_sheet(title="Graphs")

# Function to save the graph as an image
def save_graph_as_image(x, y, title, start_row):
    plt.figure(figsize=(8, 6))
    # Update the x-axis to show T1, T2, T3... instead of the actual transaction names
    x_labels = [f'T{i+1}' for i in range(len(x))]  # Create T1, T2, T3...
    plt.plot(x_labels, y, marker='o', linestyle='-', color='b')
    plt.title(title)
    plt.xlabel('Transactions')
    plt.ylabel('Variance (%)')
    plt.grid(True)
    plt.tight_layout()
    
    # Save plot to a BytesIO object
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    
    # Generate a valid Excel cell reference (e.g., "A{start_row}")
    cell_reference = f"A{start_row}"
    
    # Write the plot into the workbook
    img = Image(buf)
    img.anchor = cell_reference
    ws_graphs.add_image(img)

# Add graphs for R5 vs other reports
start_row = 1  # Starting row for the graphs in the Graphs sheet
for i in range(1, 5):
    variance_column = f'R5 Vs R{i}'
    transactions = merged_df['Transactions']
    variances = merged_df[variance_column]
    save_graph_as_image(transactions, variances, f"Graph of {variance_column}", start_row)
    start_row += 15  # Adjust row spacing for the next graph

# Save the workbook with both sheets
wb.save(output_file)

print(f"Consolidated report with updated column names, variance calculations, formatting, borders, and graphs saved to {output_file}")
