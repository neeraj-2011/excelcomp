import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font

# Set the folder containing the reports and output file name
input_folder = "./input_reports/"  # Replace with your folder path
output_file = 'consolidated_report.xlsx'

# Get the list of Excel files in the folder
excel_files = sorted(glob.glob(input_folder + "*.xlsx"))  # Ensure files are sorted by name

# Initialize a DataFrame for Report 1 to use as the base order
base_df = pd.DataFrame()

# Initialize an empty DataFrame to store the merged data
merged_df = pd.DataFrame()

# Loop through each file and merge data
for i, file in enumerate(excel_files, start=1):
    # Read the file into a DataFrame and rename columns explicitly
    df = pd.read_excel(file, header=0)  # Assuming the first row has headers
    df.columns = ['Transactions', f'Report {i} time(90%)']  # Force consistent column names
    
    if i == 1:
        # Use the first report as the base for order
        base_df = df
    
    # Merge with the main DataFrame
    if merged_df.empty:
        merged_df = df
    else:
        merged_df = pd.merge(merged_df, df, on='Transactions', how='outer')

# Sort by the base DataFrame's order
merged_df = base_df[['Transactions']].merge(merged_df, on='Transactions', how='left')

# Replace NaN with empty strings if needed
merged_df.fillna("", inplace=True)

# Reorder columns so the reports are in order (Report 1, Report 2, ..., Report 5)
columns_order = ['Transactions'] + [f'Report {i} time(90%)' for i in range(1, len(excel_files) + 1)]
merged_df = merged_df[columns_order]

# Write the merged DataFrame to an Excel file
merged_df.to_excel(output_file, index=False)

# Apply Font Color Formatting
wb = load_workbook(output_file)
ws = wb.active

# Define font colors
green_font = Font(color="00FF00")   # Green
orange_font = Font(color="FFA500") # Orange
red_font = Font(color="FF0000")    # Red

# Loop through the cells to apply formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, (int, float)):  # Ensure it's a number
            if cell.value >= 2:
                cell.font = red_font
            elif 1.8 <= cell.value < 2:
                cell.font = orange_font
            else:
                cell.font = green_font

# Save the workbook
wb.save(output_file)

print(f"Consolidated report with font color formatting saved to {output_file}")
